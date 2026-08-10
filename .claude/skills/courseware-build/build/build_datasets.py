#!/usr/bin/env python3
"""Generate mock analysis data for every activity — Excel (.xlsx, with live formulas,
industry colour-coding, Arial) AND CSV mirrors — one folder per activity (activities/activityNN/).

Fictional companies: GreenLeaf Trading Pte Ltd (activities 1-4, 6-7, 9) and
99 Agency (activity 5, matching the in-class case study). All figures are mock data
for training. Run recalc (LibreOffice) after generation to bake formula values.
"""
import os, sys, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__)); sys.path.insert(0, HERE)

def _find_repo(start):
    env = os.environ.get("COURSE_REPO")
    if env and os.path.isdir(env): return env
    d = start
    for _ in range(8):
        d = os.path.dirname(d)
        if os.path.isdir(os.path.join(d, "courseware")) and os.path.isdir(os.path.join(d, "activities")): return d
    return os.path.dirname(os.path.dirname(HERE))
REPO = _find_repo(HERE)
ACTS = os.path.join(REPO, "activities")
def _folder(name):
    num = name.split("-")[1]                     # activity-01-... -> activity01
    d = os.path.join(ACTS, f"activity{num}"); os.makedirs(d, exist_ok=True)
    return d

ARIAL = "Arial"
BOLD = Font(name=ARIAL, bold=True)
HDR_FILL = PatternFill("solid", start_color="E8F0FE")
INPUT = Font(name=ARIAL, color="0000FF")            # blue = hardcoded inputs
YELLOW = PatternFill("solid", start_color="FFFF00") # cells the learner fills
MONEY = '$#,##0;($#,##0);"-"'
PCT = "0.0%"
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def sheet_defaults(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def put(ws, r, c, v, bold=False, money=False, pct=False, inp=False, fill=None, note=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name=ARIAL, bold=bold, color="0000FF" if inp else "000000")
    if money: cell.number_format = MONEY
    if pct: cell.number_format = PCT
    if fill: cell.fill = fill
    return cell

def header_row(ws, r, values, start=1):
    for i, v in enumerate(values):
        c = put(ws, r, start + i, v, bold=True); c.fill = HDR_FILL

def write_csv(name, rows):
    p = os.path.join(_folder(name), name)
    with open(p, "w", newline="") as f:
        csv.writer(f).writerows(rows)
    print("Saved", p)

def save(wb, name):
    p = os.path.join(_folder(name), name); wb.save(p); print("Saved", p)

# ══════════════════════════════════════════ Lab 1 — financial statements
def lab01():
    wb = Workbook(); wb.remove(wb.active)
    IS = [("Revenue – Product Sales", 850000), ("Revenue – Services", 350000)]
    COGS = [("Purchases", 420000), ("Freight Inwards", 35000)]
    SGA = [("Salaries", 210000), ("Office Rent", 96000), ("Marketing", 48000),
           ("Utilities", 18000), ("Depreciation", 30000), ("Other Expenses", 22000)]
    ws = wb.create_sheet("Income Statement"); sheet_defaults(ws, [38, 16])
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — Income Statement FY2025 (S$)", bold=True)
    r = 3; header_row(ws, r, ["Line Item", "Amount"]); r += 1
    for n, v in IS: put(ws, r, 1, n); put(ws, r, 2, v, money=True, inp=True); r += 1
    put(ws, r, 1, "Total Revenue", bold=True); put(ws, r, 2, "=SUM(B4:B5)", money=True, bold=True); rev = r; r += 1
    for n, v in COGS: put(ws, r, 1, n); put(ws, r, 2, v, money=True, inp=True); r += 1
    put(ws, r, 1, "Total Cost of Goods Sold", bold=True); put(ws, r, 2, f"=SUM(B{rev+1}:B{r-1})", money=True, bold=True); cogs = r; r += 1
    put(ws, r, 1, "Gross Profit", bold=True); put(ws, r, 2, f"=B{rev}-B{cogs}", money=True, bold=True); gp = r; r += 1
    for n, v in SGA: put(ws, r, 1, n); put(ws, r, 2, v, money=True, inp=True); r += 1
    put(ws, r, 1, "Total SG&A", bold=True); put(ws, r, 2, f"=SUM(B{gp+1}:B{r-1})", money=True, bold=True); sga = r; r += 1
    put(ws, r, 1, "Operating Profit", bold=True); put(ws, r, 2, f"=B{gp}-B{sga}", money=True, bold=True); op = r; r += 1
    put(ws, r, 1, "Interest Expense"); put(ws, r, 2, 12000, money=True, inp=True); r += 1
    put(ws, r, 1, "Profit Before Tax", bold=True); put(ws, r, 2, f"=B{op}-B{r-1}", money=True, bold=True); pbt = r; r += 1
    put(ws, r, 1, "Tax (17%)"); put(ws, r, 2, f"=B{pbt}*0.17", money=True); r += 1
    put(ws, r, 1, "Net Profit", bold=True); put(ws, r, 2, f"=B{pbt}-B{r-1}", money=True, bold=True)
    put(ws, r + 2, 1, "Gross Profit Margin"); put(ws, r + 2, 2, f"=B{gp}/B{rev}", pct=True)
    put(ws, r + 3, 1, "Net Profit Margin"); put(ws, r + 3, 2, f"=B{r}/B{rev}", pct=True)

    ws = wb.create_sheet("Balance Sheet"); sheet_defaults(ws, [38, 16])
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — Balance Sheet as at 31 Dec 2025 (S$)", bold=True)
    rows = [("ASSETS", None), ("Cash and Bank", 185000), ("Accounts Receivable", 145000),
            ("Inventory", 120000), ("Total Current Assets", "=SUM(B4:B6)"),
            ("Property, Plant & Equipment", 380000), ("Accumulated Depreciation", -60000),
            ("Net PP&E", "=B8+B9"), ("TOTAL ASSETS", "=B7+B10"),
            ("LIABILITIES", None), ("Accounts Payable", 98000), ("Short-Term Loan", 50000),
            ("Total Current Liabilities", "=SUM(B13:B14)"), ("Long-Term Loan", 200000),
            ("TOTAL LIABILITIES", "=B15+B16"),
            ("EQUITY", None), ("Share Capital", 300000), ("Retained Earnings", 122000),
            ("TOTAL EQUITY", "=SUM(B19:B20)"),
            ("Balance Check (Assets − Liabilities − Equity)", "=B11-B17-B21")]
    r = 3
    for n, v in rows:
        hd = v is None or (isinstance(v, str) and n.isupper()) or n.startswith("Total") or n.startswith("Net ") or n.startswith("Balance")
        put(ws, r, 1, n, bold=hd)
        if v is not None:
            put(ws, r, 2, v, money=True, bold=hd, inp=not isinstance(v, str))
        r += 1

    ws = wb.create_sheet("Cash Flow"); sheet_defaults(ws, [38, 16])
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — Cash Flow Statement FY2025 (S$)", bold=True)
    rows = [("OPERATING ACTIVITIES", None), ("Net Profit", 256470), ("Add: Depreciation", 30000),
            ("(Increase) in Accounts Receivable", -25000), ("(Increase) in Inventory", -15000),
            ("Increase in Accounts Payable", 8000), ("Net Cash from Operating", "=SUM(B4:B8)"),
            ("INVESTING ACTIVITIES", None), ("Purchase of Equipment (CAPEX)", -80000),
            ("Net Cash from Investing", "=B11"),
            ("FINANCING ACTIVITIES", None), ("Loan Repayment", -40000), ("Dividends Paid", -30000),
            ("Net Cash from Financing", "=SUM(B14:B15)"),
            ("Net Change in Cash", "=B9+B12+B16"), ("Opening Cash", 80530),
            ("Closing Cash", "=B17+B18")]
    r = 3
    for n, v in rows:
        hd = v is None or n.startswith("Net Cash") or n.startswith("Net Change") or n.startswith("Closing")
        put(ws, r, 1, n, bold=hd)
        if v is not None: put(ws, r, 2, v, money=True, bold=hd, inp=not isinstance(v, str))
        r += 1
    save(wb, "activity-01-financial-statements.xlsx")

    write_csv("activity-01-income-statement.csv", [["Line Item", "Amount (S$)"],
        ["Revenue - Product Sales", 850000], ["Revenue - Services", 350000], ["Total Revenue", 1200000],
        ["Purchases", 420000], ["Freight Inwards", 35000], ["Total Cost of Goods Sold", 455000],
        ["Gross Profit", 745000], ["Salaries", 210000], ["Office Rent", 96000], ["Marketing", 48000],
        ["Utilities", 18000], ["Depreciation", 30000], ["Other Expenses", 22000], ["Total SG&A", 424000],
        ["Operating Profit", 321000], ["Interest Expense", 12000], ["Profit Before Tax", 309000],
        ["Tax (17%)", 52530], ["Net Profit", 256470]])
    write_csv("activity-01-balance-sheet.csv", [["Line Item", "Amount (S$)"],
        ["Cash and Bank", 185000], ["Accounts Receivable", 145000], ["Inventory", 120000],
        ["Total Current Assets", 450000], ["Property Plant & Equipment", 380000],
        ["Accumulated Depreciation", -60000], ["Net PP&E", 320000], ["TOTAL ASSETS", 770000],
        ["Accounts Payable", 98000], ["Short-Term Loan", 50000], ["Total Current Liabilities", 148000],
        ["Long-Term Loan", 200000], ["TOTAL LIABILITIES", 348000], ["Share Capital", 300000],
        ["Retained Earnings", 122000], ["TOTAL EQUITY", 422000]])
    write_csv("activity-01-cash-flow.csv", [["Line Item", "Amount (S$)"],
        ["Net Profit", 256470], ["Add: Depreciation", 30000], ["(Increase) in Accounts Receivable", -25000],
        ["(Increase) in Inventory", -15000], ["Increase in Accounts Payable", 8000],
        ["Net Cash from Operating", 254470], ["Purchase of Equipment (CAPEX)", -80000],
        ["Net Cash from Investing", -80000], ["Loan Repayment", -40000], ["Dividends Paid", -30000],
        ["Net Cash from Financing", -70000], ["Net Change in Cash", 104470],
        ["Opening Cash", 80530], ["Closing Cash", 185000]])

# ══════════════════════════════════════════ Lab 2 — 12-month budget plan
def lab02_sales():
    s = [90000]
    for _ in range(11): s.append(round(s[-1] * 1.02))
    return s

def lab02():
    sales = lab02_sales()
    wb = Workbook(); ws = wb.active; ws.title = "Budget FY2026"
    sheet_defaults(ws, [24] + [11] * 13)
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — FY2026 Budget Plan (S$)  ·  enter these figures into Xero Budget Manager", bold=True)
    put(ws, 2, 1, "Rates:", bold=True); put(ws, 2, 2, "Commission"); put(ws, 2, 3, 0.05, inp=True, pct=True)
    put(ws, 2, 4, "Discount"); put(ws, 2, 5, 0.02, inp=True, pct=True)
    put(ws, 2, 6, "Direct Costs"); put(ws, 2, 7, 0.55, inp=True, pct=True)
    header_row(ws, 4, ["Account"] + MONTHS + ["Total"])
    fixed = [("Office Rent", 8000), ("Salaries", 17500), ("Marketing", 4000),
             ("Utilities", 1500), ("Depreciation", 2500), ("Other Expenses", 1800)]
    r = 5
    put(ws, r, 1, "Sales")
    for i, v in enumerate(sales): put(ws, r, 2 + i, v, money=True, inp=True)
    put(ws, r, 14, "=SUM(B5:M5)", money=True, bold=True); r += 1
    for name, rate_cell in [("Sales Commission", "$C$2"), ("Sales Discount", "$E$2"), ("Direct Costs", "$G$2")]:
        put(ws, r, 1, name)
        for i in range(12):
            col = get_column_letter(2 + i)
            put(ws, r, 2 + i, f"={col}5*{rate_cell}", money=True)
        put(ws, r, 14, f"=SUM(B{r}:M{r})", money=True, bold=True); r += 1
    for name, v in fixed:
        put(ws, r, 1, name)
        for i in range(12): put(ws, r, 2 + i, v, money=True, inp=True)
        put(ws, r, 14, f"=SUM(B{r}:M{r})", money=True, bold=True); r += 1
    put(ws, r, 1, "Net Surplus", bold=True)
    for i in range(12):
        col = get_column_letter(2 + i)
        put(ws, r, 2 + i, f"={col}5-SUM({col}6:{col}{r-1})", money=True, bold=True)
    put(ws, r, 14, f"=SUM(B{r}:M{r})", money=True, bold=True)
    save(wb, "activity-02-budget-plan.xlsx")

    rows = [["Account"] + MONTHS + ["Total"]]
    rows.append(["Sales"] + sales + [sum(sales)])
    for name, rate in [("Sales Commission", .05), ("Sales Discount", .02), ("Direct Costs", .55)]:
        vals = [round(x * rate) for x in sales]; rows.append([name] + vals + [sum(vals)])
    for name, v in [("Office Rent", 8000), ("Salaries", 17500), ("Marketing", 4000),
                    ("Utilities", 1500), ("Depreciation", 2500), ("Other Expenses", 1800)]:
        rows.append([name] + [v] * 12 + [v * 12])
    exp = [sum(rows[i][j] for i in range(2, 11)) for j in range(1, 14)]
    rows.append(["Net Surplus"] + [rows[1][j] - exp[j - 1] for j in range(1, 14)])
    write_csv("activity-02-budget-plan.csv", rows)

# ══════════════════════════════════════════ Lab 3 — classification worksheet
L3_ITEMS = [
    ("Monthly office rental", 96000), ("Raw material purchases", 420000),
    ("Staff salaries", 210000), ("Annual digital marketing campaign", 48000),
    ("New delivery van purchase", 85000), ("New POS system implementation", 40000),
    ("Cash float for retail outlets", 25000), ("Bank loan repayment", 40000),
    ("Utilities (electricity, water, phone)", 18000), ("Office renovation", 60000),
    ("Dividend payout to shareholders", 30000), ("Inventory build-up for year-end sale", 75000),
]
def lab03():
    wb = Workbook(); ws = wb.active; ws.title = "Worksheet"
    sheet_defaults(ws, [40, 16, 22, 24])
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — Budget Classification Worksheet", bold=True)
    put(ws, 2, 1, "Classify every item: Budget Type = Operating / Cash / Capital / Financial · Method = Baseline / Incremental / Zero-based / Hybrid")
    header_row(ws, 4, ["Line Item", "Annual Amount", "Budget Type", "Preparation Method"])
    for i, (n, v) in enumerate(L3_ITEMS):
        r = 5 + i
        put(ws, r, 1, n); put(ws, r, 2, v, money=True, inp=True)
        put(ws, r, 3, "", fill=YELLOW); put(ws, r, 4, "", fill=YELLOW)
    ref = wb.create_sheet("Reference"); sheet_defaults(ref, [22, 80])
    defs = [("Operating budget", "Revenue and day-to-day operating expenses."),
            ("Cash budget", "Cash inflows and outflows — liquidity planning."),
            ("Capital budget", "Long-term investments: plant, equipment, systems, renovation."),
            ("Financial budget", "Funding, loans, repayments and shareholder distributions."),
            ("Baseline", "Begins with the previous plan as the baseline."),
            ("Incremental", "A % or $ increment on the previous baseline."),
            ("Zero-based", "Starts fresh like a new plan."),
            ("Hybrid", "A combination of the above methods.")]
    header_row(ref, 1, ["Term", "Definition"])
    for i, (t, d) in enumerate(defs): put(ref, 2 + i, 1, t, bold=True); put(ref, 2 + i, 2, d)
    save(wb, "activity-03-budget-classification.xlsx")
    write_csv("activity-03-budget-classification.csv",
              [["Line Item", "Annual Amount (S$)", "Budget Type", "Preparation Method"]] +
              [[n, v, "", ""] for n, v in L3_ITEMS])

# ══════════════════════════════════════════ Lab 4 — FY2025 actuals baseline
def lab04_rows():
    sales = [75000]
    for _ in range(11): sales.append(round(sales[-1] * 1.015))
    rows = [["Sales"] + sales,
            ["Sales Commission (5%)"] + [round(x * .05) for x in sales],
            ["Sales Discount (2%)"] + [round(x * .02) for x in sales],
            ["Direct Costs (55%)"] + [round(x * .55) for x in sales],
            ["Office Rent"] + [7500] * 12,
            ["Salaries (10 staff)"] + [45000] * 12,
            ["Depreciation"] + [12000] * 12,
            ["Other Expenses"] + [5000] * 12]
    return rows

def lab04():
    rows = lab04_rows()
    wb = Workbook(); ws = wb.active; ws.title = "FY2025 Actuals"
    sheet_defaults(ws, [24] + [11] * 13)
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — FY2025 Monthly Actuals (S$) — baseline for the FY2026 forecast", bold=True)
    header_row(ws, 3, ["Account"] + MONTHS + ["Total"])
    for i, row in enumerate(rows):
        r = 4 + i
        put(ws, r, 1, row[0])
        for j, v in enumerate(row[1:]): put(ws, r, 2 + j, v, money=True, inp=True)
        put(ws, r, 14, f"=SUM(B{r}:M{r})", money=True, bold=True)
    a = wb.create_sheet("Forecast Assumptions"); sheet_defaults(a, [46, 14])
    put(a, 1, 1, "FY2026 Forecast Assumptions — apply these in Xero Budget Manager / the Forecast", bold=True)
    ass = [("Sales growth vs same month last year", 0.10, PCT),
           ("Sales Commission (% of Sales)", 0.05, PCT), ("Sales Discount (% of Sales)", 0.02, PCT),
           ("Direct Costs (% of Revenue)", 0.55, PCT),
           ("Additional office from June (x current rent)", 1.5, "0.0"),
           ("Additional full-time headcount from July", 3, "0"),
           ("Average salary per head (per month)", 4500, MONEY),
           ("Depreciation on new capex (per month, from Jan)", 30000, MONEY),
           ("Income tax", "Ignore", None)]
    header_row(a, 3, ["Assumption", "Value"])
    for i, (n, v, fmt) in enumerate(ass):
        r = 4 + i
        put(a, r, 1, n); c = put(a, r, 2, v, inp=True); c.fill = YELLOW
        if fmt == PCT: c.number_format = PCT
        elif fmt: c.number_format = fmt
    save(wb, "activity-04-fy2025-actuals.xlsx")
    csv_rows = [["Account"] + MONTHS + ["Total"]] + [r + [sum(r[1:])] for r in [list(x) for x in rows]]
    write_csv("activity-04-fy2025-actuals.csv", csv_rows)

# ══════════════════════════════════════════ Lab 5 — 99 Agency template
L5_CLIENTS = [("Alpha Media", 320000, 0.12), ("Beta Retail", 280000, 0.08),
              ("City Schools", 240000, 0.15), ("Delta Tech", 200000, 0.10),
              ("Other Clients", 360000, 0.05)]
def lab05():
    wb = Workbook(); wb.remove(wb.active)
    a = wb.create_sheet("Assumptions"); sheet_defaults(a, [46, 16, 14])
    put(a, 1, 1, "99 Agency — FY2026 Master Budget Working Template — Key Assumptions", bold=True)
    rows = [("Market size (addressable, S$)", 12000000, MONEY), ("99 Agency market share", 0.12, PCT),
            ("Committee rule", "Average of bottom-up and top-down", None),
            ("Ad Campaigns cost of sales (% of Ad revenue)", 0.30, PCT),
            ("SEO Campaigns cost of sales (% of SEO revenue)", 0.25, PCT),
            ("Revenue split — Ad : SEO", "60% : 40%", None),
            ("Fixed SG&A: Office rent (annual)", 60000, MONEY), ("Fixed SG&A: Accounting services", 12000, MONEY),
            ("Fixed SG&A: Legal fees", 8000, MONEY), ("Fixed SG&A: Staff training", 10000, MONEY),
            ("Fixed SG&A: Management salaries", 180000, MONEY), ("Fixed SG&A: Administration", 24000, MONEY),
            ("Variable SG&A: External services (% revenue)", 0.04, PCT),
            ("Variable SG&A: Selling commissions (% revenue)", 0.03, PCT),
            ("CAPEX Q1 / Q2 / Q3 / Q4", "40,000 / 60,000 / 25,000 / 80,000", None),
            ("Beginning Fixed Assets", 400000, MONEY), ("Depreciation (% of beginning Fixed Assets)", 0.10, PCT),
            ("Financial liabilities: beginning", 3200000, MONEY), ("Financial liabilities: ending (target)", 2800000, MONEY),
            ("Interest rate", 0.05, PCT), ("Accounts receivable days", 45, "0"), ("Accounts payable days", 30, "0")]
    header_row(a, 3, ["Assumption", "Value"])
    for i, (n, v, fmt) in enumerate(rows):
        r = 4 + i; put(a, r, 1, n); c = put(a, r, 2, v, inp=True)
        if fmt: c.number_format = fmt

    rv = wb.create_sheet("1 Revenue"); sheet_defaults(rv, [24, 16, 12, 18])
    put(rv, 1, 1, "STEP 1 — Revenue Budget (bottom-up per client, then compare with top-down)", bold=True)
    header_row(rv, 3, ["Client", "FY2025 Revenue", "Growth", "FY2026 Bottom-Up"])
    r = 4
    for n, v, g in L5_CLIENTS:
        put(rv, r, 1, n); put(rv, r, 2, v, money=True, inp=True); put(rv, r, 3, g, pct=True, inp=True)
        put(rv, r, 4, f"=B{r}*(1+C{r})", money=True); r += 1
    put(rv, r, 1, "Bottom-up total", bold=True); put(rv, r, 4, f"=SUM(D4:D{r-1})", money=True, bold=True)
    put(rv, r + 1, 1, "Top-down target (market × share)", bold=True)
    put(rv, r + 1, 4, "=Assumptions!B4*Assumptions!B5", money=True, bold=True)
    put(rv, r + 2, 1, "Committee revenue target (average)", bold=True)
    c = put(rv, r + 2, 4, f"=AVERAGE(D{r},D{r+1})", money=True, bold=True); c.fill = YELLOW

    cs = wb.create_sheet("2 Cost of Sales"); sheet_defaults(cs, [34, 18])
    put(cs, 1, 1, "STEP 2 — Cost of Sales Budget (baseline + averaging)", bold=True)
    header_row(cs, 3, ["Line", "FY2026"])
    put(cs, 4, 1, "Ad Campaigns revenue (60%)"); put(cs, 4, 2, "='1 Revenue'!D11*0.6", money=True)
    put(cs, 5, 1, "SEO Campaigns revenue (40%)"); put(cs, 5, 2, "='1 Revenue'!D11*0.4", money=True)
    put(cs, 6, 1, "Ad Campaigns cost (30%)"); put(cs, 6, 2, "=B4*Assumptions!B7", money=True)
    put(cs, 7, 1, "SEO Campaigns cost (25%)"); put(cs, 7, 2, "=B5*Assumptions!B8", money=True)
    put(cs, 8, 1, "Total Cost of Sales", bold=True); put(cs, 8, 2, "=B6+B7", money=True, bold=True)

    for name, title in [("3 SGA", "STEP 3 — SG&A Budget"), ("4 Fixed Assets", "STEP 4 — Fixed Assets & Depreciation"),
                        ("5 Working Capital", "STEP 5 — Working Capital"), ("6 Liabilities", "STEP 6 — Financial Liabilities"),
                        ("7 Master Budget", "STEP 7 — Master Budget Compilation (Income Statement · Balance Sheet · Cash Flow)")]:
        ws = wb.create_sheet(name); sheet_defaults(ws, [36, 16, 16, 16, 16])
        put(ws, 1, 1, title, bold=True)
        put(ws, 2, 1, "Build this schedule from the Assumptions sheet — follow Activity 5 in the Learner Guide.", fill=YELLOW)
    save(wb, "activity-05-99-agency-template.xlsx")

    write_csv("activity-05-client-baseline.csv",
              [["Client", "FY2025 Revenue (S$)", "Expected FY2026 Growth"]] +
              [[n, v, g] for n, v, g in L5_CLIENTS])
    write_csv("activity-05-assumptions.csv",
              [["Assumption", "Value"]] + [[n, v] for n, v, _ in rows])

# ══════════════════════════════════════════ Lab 6 — budget control monitor
L6 = [("Operations", 480000, 240000, 262000), ("Sales", 300000, 150000, 141000),
      ("Marketing", 120000, 60000, 71500), ("Human Resources", 90000, 45000, 44200),
      ("Finance", 75000, 37500, 36900), ("IT", 110000, 55000, 49800)]
def lab06():
    wb = Workbook(); ws = wb.active; ws.title = "Control Monitor"
    sheet_defaults(ws, [20, 16, 14, 14, 14, 10, 12])
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — Mid-Year Budget Control Monitor (S$)", bold=True)
    put(ws, 2, 1, "Variance threshold:"); put(ws, 2, 2, 0.10, inp=True, pct=True, fill=YELLOW)
    header_row(ws, 4, ["Department", "Annual Budget", "YTD Budget", "YTD Actual", "Variance", "Var %", "Status"])
    for i, (d, ann, ytdb, ytda) in enumerate(L6):
        r = 5 + i
        put(ws, r, 1, d); put(ws, r, 2, ann, money=True, inp=True)
        put(ws, r, 3, ytdb, money=True, inp=True); put(ws, r, 4, ytda, money=True, inp=True)
        put(ws, r, 5, f"=D{r}-C{r}", money=True); put(ws, r, 6, f"=E{r}/C{r}", pct=True)
        put(ws, r, 7, f'=IF(ABS(F{r})>$B$2,"REVIEW","OK")')
    r = 5 + len(L6)
    put(ws, r, 1, "Total", bold=True)
    for col in "BCDE": put(ws, r, ord(col) - 64, f"=SUM({col}5:{col}{r-1})", money=True, bold=True)
    save(wb, "activity-06-budget-control-monitor.xlsx")
    write_csv("activity-06-budget-control-monitor.csv",
              [["Department", "Annual Budget (S$)", "YTD Budget (S$)", "YTD Actual (S$)"]] +
              [[d, a, b, c] for d, a, b, c in L6])

# ══════════════════════════════════════════ Lab 7 — budget vs actual
L7_ACCOUNTS = [("Sales", "income", 90000, 1.02, 0.97),
               ("Sales Commission", "expense", 4500, 1.02, 0.97),
               ("Direct Costs", "expense", 49500, 1.02, 1.04),
               ("Office Rent", "expense", 8000, 1.0, 1.0),
               ("Salaries", "expense", 17500, 1.0, 1.02),
               ("Marketing", "expense", 4000, 1.0, 1.28),
               ("Utilities", "expense", 1500, 1.0, 1.07),
               ("Other Expenses", "expense", 1800, 1.0, 0.92)]
def l7_series(base, g):
    out = [base]
    for _ in range(11): out.append(round(out[-1] * g))
    return out
def lab07():
    wb = Workbook(); wb.remove(wb.active)
    budget = {n: l7_series(b, g) for n, t, b, g, f in L7_ACCOUNTS}
    actual = {n: [round(v * f) for v in budget[n]] for n, t, b, g, f in L7_ACCOUNTS}
    for sheet, data in [("Budget", budget), ("Actual", actual)]:
        ws = wb.create_sheet(sheet); sheet_defaults(ws, [22] + [11] * 13)
        put(ws, 1, 1, f"GreenLeaf Trading Pte Ltd — FY2026 {sheet} (S$)", bold=True)
        header_row(ws, 3, ["Account"] + MONTHS + ["Total"])
        for i, (n, t, b, g, f) in enumerate(L7_ACCOUNTS):
            r = 4 + i; put(ws, r, 1, n)
            for j, v in enumerate(data[n]): put(ws, r, 2 + j, v, money=True, inp=True)
            put(ws, r, 14, f"=SUM(B{r}:M{r})", money=True, bold=True)
    ws = wb.create_sheet("Variance"); sheet_defaults(ws, [22, 12, 14, 14, 14, 10, 14])
    put(ws, 1, 1, "Variance Analysis — Actual vs Budget (full year)", bold=True)
    put(ws, 2, 1, "Threshold:"); put(ws, 2, 2, 0.10, inp=True, pct=True, fill=YELLOW)
    header_row(ws, 4, ["Account", "Type", "Budget", "Actual", "Variance", "Var %", "Classification"])
    for i, (n, t, b, g, f) in enumerate(L7_ACCOUNTS):
        r = 5 + i
        put(ws, r, 1, n); put(ws, r, 2, t)
        put(ws, r, 3, f"=Budget!N{4+i}", money=True); put(ws, r, 4, f"=Actual!N{4+i}", money=True)
        put(ws, r, 5, f"=D{r}-C{r}", money=True); put(ws, r, 6, f"=E{r}/C{r}", pct=True)
        put(ws, r, 7, f'=IF(B{r}="income",IF(E{r}>=0,"Favourable","Adverse"),IF(E{r}<=0,"Favourable","Adverse"))')
    save(wb, "activity-07-budget-vs-actual.xlsx")
    for name, data in [("activity-07-budget.csv", budget), ("activity-07-actual.csv", actual)]:
        rows = [["Account"] + MONTHS + ["Total"]]
        for n, t, b, g, f in L7_ACCOUNTS:
            rows.append([n] + data[n] + [sum(data[n])])
        write_csv(name, rows)

# ══════════════════════════════════════════ Lab 8 — Xero-style exports for BI
def lab08():
    contacts = ["Acme Retail", "Bright Media", "Coastal Foods", "Delta Logistics", "Evergreen Mart",
                "Fusion Events", "Galaxy Tech", "Harbour Cafe", "Island Traders", "Jade Wellness"]
    invoices, bills = [], []
    inv_no = 1001
    for m in range(1, 7):
        for i, c in enumerate(contacts):
            amt = 4000 + (i * 850) + (m * 320)
            status = "Paid" if (i + m) % 3 else "Authorised"
            due_m = m + 1
            invoices.append([f"INV-{inv_no}", f"2026-{m:02d}-{5 + i*2:02d}", f"2026-{due_m:02d}-{5 + i*2:02d}",
                             c, amt, amt if status == "Paid" else 0, status])
            inv_no += 1
    suppliers = ["Media Buy Pte Ltd", "CloudHost SG", "Office Supplies Co", "Freelance Design",
                 "Utilities Board", "Landlord Holdings", "Insurance Direct", "Logistics Partner"]
    bill_no = 501
    for m in range(1, 7):
        for i, s in enumerate(suppliers):
            amt = 1500 + (i * 620) + (m * 180)
            status = "Paid" if (i + m) % 4 else "Authorised"
            bills.append([f"BILL-{bill_no}", f"2026-{m:02d}-{3 + i*3:02d}", f"2026-{m+1:02d}-{3 + i*3:02d}",
                          s, amt, amt if status == "Paid" else 0, status])
            bill_no += 1
    tb = [("Cash and Bank", 185000, 0), ("Accounts Receivable", 145000, 0), ("Inventory", 120000, 0),
          ("Property, Plant & Equipment", 380000, 0), ("Accumulated Depreciation", 0, 60000),
          ("Accounts Payable", 0, 98000), ("Short-Term Loan", 0, 50000), ("Long-Term Loan", 0, 200000),
          ("Share Capital", 0, 300000), ("Retained Earnings", 0, 122000),
          ("Sales", 0, 612000), ("Cost of Goods Sold", 336600, 0), ("Salaries", 105000, 0),
          ("Rent", 48000, 0), ("Marketing", 24000, 0), ("Utilities", 9000, 0),
          ("Depreciation Expense", 15000, 0), ("Other Expenses", 11000, 0), ("Interest Expense", 6000, 0),
          ("Suspense (balancing)", 57400, 0)]
    wb = Workbook(); wb.remove(wb.active)
    for sheet, hdr, data in [("Invoices", ["Invoice No", "Date", "Due Date", "Contact", "Amount", "Amount Paid", "Status"], invoices),
                             ("Bills", ["Bill No", "Date", "Due Date", "Supplier", "Amount", "Amount Paid", "Status"], bills)]:
        ws = wb.create_sheet(sheet); sheet_defaults(ws, [12, 12, 12, 22, 12, 12, 12])
        header_row(ws, 1, hdr)
        for i, row in enumerate(data):
            for j, v in enumerate(row):
                put(ws, 2 + i, 1 + j, v, money=(j in (4, 5)), inp=True)
        n = len(data) + 1
        put(ws, n + 1, 4, "Total", bold=True); put(ws, n + 1, 5, f"=SUM(E2:E{n})", money=True, bold=True)
        put(ws, n + 1, 6, f"=SUM(F2:F{n})", money=True, bold=True)
        put(ws, n + 2, 4, "Outstanding", bold=True); put(ws, n + 2, 5, f"=E{n+1}-F{n+1}", money=True, bold=True)
    ws = wb.create_sheet("Trial Balance"); sheet_defaults(ws, [30, 14, 14])
    header_row(ws, 1, ["Account", "Debit", "Credit"])
    for i, (n, d, c) in enumerate(tb):
        put(ws, 2 + i, 1, n); put(ws, 2 + i, 2, d, money=True, inp=True); put(ws, 2 + i, 3, c, money=True, inp=True)
    r = len(tb) + 2
    put(ws, r, 1, "Total", bold=True)
    put(ws, r, 2, f"=SUM(B2:B{r-1})", money=True, bold=True); put(ws, r, 3, f"=SUM(C2:C{r-1})", money=True, bold=True)
    put(ws, r + 1, 1, "Balance check (Debit − Credit)"); put(ws, r + 1, 2, f"=B{r}-C{r}", money=True)
    save(wb, "activity-08-xero-export.xlsx")
    write_csv("activity-08-invoices.csv", [["Invoice No", "Date", "Due Date", "Contact", "Amount", "Amount Paid", "Status"]] + invoices)
    write_csv("activity-08-bills.csv", [["Bill No", "Date", "Due Date", "Supplier", "Amount", "Amount Paid", "Status"]] + bills)
    write_csv("activity-08-trial-balance.csv", [["Account", "Debit", "Credit"]] + [[n, d, c] for n, d, c in tb])

# ══════════════════════════════════════════ Lab 9 — approval pack
L9 = [("Operations", 452000, 495000, "Two new production lines ramp up"),
      ("Sales", 287000, 340000, "Expansion into two new regions"),
      ("Marketing", 118000, 155000, "Brand refresh + always-on digital"),
      ("Human Resources", 88000, 92000, "Statutory increments only"),
      ("Finance", 74000, 76000, "Audit fee increase"),
      ("IT", 105000, 150000, "ERP budgeting module rollout")]
def lab09():
    wb = Workbook(); ws = wb.active; ws.title = "Submissions"
    sheet_defaults(ws, [20, 16, 18, 12, 44])
    put(ws, 1, 1, "GreenLeaf Trading Pte Ltd — FY2027 Departmental Budget Submissions (S$)", bold=True)
    put(ws, 2, 1, "HQ growth cap (total ask vs prior year):"); put(ws, 2, 2, 0.08, inp=True, pct=True, fill=YELLOW)
    header_row(ws, 4, ["Department", "FY2026 Actual", "FY2027 Requested", "Growth %", "Justification"])
    for i, (d, prev, req, j) in enumerate(L9):
        r = 5 + i
        put(ws, r, 1, d); put(ws, r, 2, prev, money=True, inp=True); put(ws, r, 3, req, money=True, inp=True)
        put(ws, r, 4, f"=(C{r}-B{r})/B{r}", pct=True); put(ws, r, 5, j)
    r = 5 + len(L9)
    put(ws, r, 1, "Total", bold=True)
    put(ws, r, 2, f"=SUM(B5:B{r-1})", money=True, bold=True); put(ws, r, 3, f"=SUM(C5:C{r-1})", money=True, bold=True)
    put(ws, r, 4, f"=(C{r}-B{r})/B{r}", pct=True)
    put(ws, r + 1, 1, "Within HQ cap?"); put(ws, r + 1, 2, f'=IF(D{r}<=B2,"YES","NO — committee must trim")')
    save(wb, "activity-09-approval-pack.xlsx")
    write_csv("activity-09-approval-pack.csv",
              [["Department", "FY2026 Actual (S$)", "FY2027 Requested (S$)", "Justification"]] +
              [[d, p, q, j] for d, p, q, j in L9])

# ══════════════════════════════════════════ Labs 10-12 — tax worksheets
L10 = [("1 Jan 2019", "31 Dec 2020"), ("1 Apr 2019", "31 Mar 2020"), ("1 Jul 2019", "30 Jun 2020"),
       ("1 Jan 2025", "31 Dec 2025"), ("1 Oct 2024", "30 Sep 2025")]
def lab10():
    wb = Workbook(); ws = wb.active; ws.title = "YA Worksheet"
    sheet_defaults(ws, [22, 22, 22])
    put(ws, 1, 1, "Year of Assessment Worksheet — YA = the year following the financial year end", bold=True)
    header_row(ws, 3, ["Financial Year Start", "Financial Year End", "Year of Assessment (YA)"])
    for i, (s, e) in enumerate(L10):
        r = 4 + i; put(ws, r, 1, s); put(ws, r, 2, e); put(ws, r, 3, "", fill=YELLOW)
    save(wb, "activity-10-ya-worksheet.xlsx")
    write_csv("activity-10-ya-worksheet.csv",
              [["Financial Year Start", "Financial Year End", "Year of Assessment (YA)"]] +
              [[s, e, ""] for s, e in L10])

def lab11():
    wb = Workbook(); ws = wb.active; ws.title = "Filing Deadlines"
    sheet_defaults(ws, [22, 22, 26, 26])
    put(ws, 1, 1, "ECI & Corporate Tax Filing Deadlines — ECI within 3 months of FY end · e-File corporate tax by 30 Nov", bold=True)
    header_row(ws, 3, ["Financial Year Start", "Financial Year End", "ECI Filing Deadline", "Corporate Tax e-Filing Deadline"])
    for i, (s, e) in enumerate(L10):
        r = 4 + i; put(ws, r, 1, s); put(ws, r, 2, e)
        put(ws, r, 3, "", fill=YELLOW); put(ws, r, 4, "", fill=YELLOW)
    save(wb, "activity-11-filing-deadlines.xlsx")
    write_csv("activity-11-filing-deadlines.csv",
              [["Financial Year Start", "Financial Year End", "ECI Filing Deadline", "Corporate Tax e-Filing Deadline"]] +
              [[s, e, "", ""] for s, e in L10])

L12 = [(1, 200000), (2, 300000), (3, 400000), (4, 500000), (5, 600000)]
def lab12():
    wb = Workbook(); ws = wb.active; ws.title = "Tax Computation"
    sheet_defaults(ws, [10, 14, 18, 22, 14])
    put(ws, 1, 1, "Start-Up Tax Computation (S$) — corporate tax rate 17%", bold=True)
    header_row(ws, 3, ["Year", "Profits", "Exempt Amount", "Chargeable After Exemption", "Tax @ 17%"])
    for i, (y, p) in enumerate(L12):
        r = 4 + i
        put(ws, r, 1, f"Year {y}"); put(ws, r, 2, p, money=True, inp=True)
        put(ws, r, 3, "", fill=YELLOW); put(ws, r, 4, "", fill=YELLOW); put(ws, r, 5, "", fill=YELLOW)
    ref = wb.create_sheet("Rates"); sheet_defaults(ref, [56, 16])
    put(ref, 1, 1, "Exemption rules (from YA 2020)", bold=True)
    rules = [("Corporate tax rate (flat, on chargeable income)", "17%"),
             ("Start-Up Tax Exemption — first 3 YAs: exemption on first S$100,000", "75%"),
             ("Start-Up Tax Exemption — first 3 YAs: exemption on next S$100,000", "50%"),
             ("Partial Tax Exemption — from YA 4: exemption on first S$10,000", "75%"),
             ("Partial Tax Exemption — from YA 4: exemption on next S$190,000", "50%")]
    header_row(ref, 3, ["Rule", "Rate"])
    for i, (n, v) in enumerate(rules): put(ref, 4 + i, 1, n); put(ref, 4 + i, 2, v, inp=True)
    save(wb, "activity-12-startup-tax.xlsx")
    write_csv("activity-12-startup-tax.csv",
              [["Year", "Profits (S$)", "Exempt Amount (S$)", "Chargeable After Exemption (S$)", "Tax @ 17% (S$)"]] +
              [[f"Year {y}", p, "", "", ""] for y, p in L12])

if __name__ == "__main__":
    for fn in (lab01, lab02, lab03, lab04, lab05, lab06, lab07, lab08, lab09, lab10, lab11, lab12):
        fn()
    print("All activity datasets generated in activities/")
